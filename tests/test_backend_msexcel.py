import logging
from io import BytesIO
from pathlib import Path

import pytest
from docling_core.transforms.serializer.markdown import MarkdownParams
from docling_core.transforms.serializer.markdown_excel import (
    MsExcelMarkdownDocSerializer,
)
from docling_core.types.doc import (
    ContentLayer,
    GroupLabel,
    PictureClassificationLabel,
    PictureItem,
    TableItem,
    TextItem,
)
from docling_core.types.doc.document import DEFAULT_CONTENT_LAYERS
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment

from docling.backend.docx.drawingml.utils import get_libreoffice_cmd
from docling.backend.msexcel_backend import (
    ExcelCell,
    ExcelTable,
    MsExcelDocumentBackend,
)
from docling.datamodel.backend_options import MsExcelBackendOptions
from docling.datamodel.base_models import InputFormat
from docling.datamodel.document import ConversionResult, DoclingDocument, InputDocument
from docling.document_converter import DocumentConverter, ExcelFormatOption

from .test_data_gen_flag import GEN_TEST_DATA
from .verify_utils import verify_document, verify_export

_log = logging.getLogger(__name__)

GENERATE = GEN_TEST_DATA


@pytest.fixture(scope="module")
def libreoffice_available() -> bool:
    """Return True when a working LibreOffice installation is detected."""
    try:
        return get_libreoffice_cmd(raise_if_unavailable=True) is not None
    except Exception:
        return False


def get_excel_paths():
    # Define the directory you want to search
    directory = Path("./tests/data/xlsx/sources/")

    # List all Excel files in the directory and its subdirectories.
    # Exclude ~$ prefixed lock files created by Excel when a file is open.
    excel_files = sorted(
        f for f in directory.rglob("*.xlsx") if not f.name.startswith("~$")
    ) + sorted(f for f in directory.rglob("*.xlsm") if not f.name.startswith("~$"))
    return excel_files


def get_converter():
    converter = DocumentConverter(allowed_formats=[InputFormat.XLSX])

    return converter


@pytest.fixture(scope="module")
def documents() -> list[tuple[Path, DoclingDocument]]:
    documents: list[dict[Path, DoclingDocument]] = []

    excel_paths = get_excel_paths()
    converter = get_converter()

    for excel_path in excel_paths:
        _log.debug(f"converting {excel_path}")

        gt_path = excel_path.parent.parent / "groundtruth" / excel_path.name

        conv_result: ConversionResult = converter.convert(excel_path)

        doc: DoclingDocument = conv_result.document

        assert doc, f"Failed to convert document from file {gt_path}"
        documents.append((gt_path, doc))

    return documents


def test_comments_extraction(documents) -> None:
    """Test that cell comments are extracted into the NOTES content layer."""
    from docling_core.types.doc import GroupItem

    doc = next(item for path, item in documents if path.stem == "xlsx_comments")

    comment_groups = [
        g
        for g in doc.groups
        if isinstance(g, GroupItem) and g.name.startswith("comment-")
    ]
    assert len(comment_groups) == 4, (
        f"Expected 4 comment groups (2 notes + 2 threaded), got {len(comment_groups)}"
    )

    comment_texts = [
        t.text
        for t in doc.texts
        if isinstance(t, TextItem) and t.content_layer == ContentLayer.NOTES
    ]

    # Check for old-style notes
    assert any("John Reviewer" in t for t in comment_texts), (
        "Expected 'John Reviewer' in comment texts"
    )
    assert any("Jane Editor" in t for t in comment_texts), (
        "Expected 'Jane Editor' in comment texts"
    )
    assert any("Why Python" in t for t in comment_texts), (
        "Expected comment body text content"
    )

    # Check for threaded comments with author and timestamp
    assert any("Marcus Sterling" in t and "time:" in t for t in comment_texts), (
        "Expected threaded comment with author Marcus Sterling and timestamp"
    )
    assert any("Jane Smith" in t and "time:" in t for t in comment_texts), (
        "Expected threaded comment with author Jane Smith and timestamp"
    )
    assert any("never thought it would be so low" in t for t in comment_texts), (
        "Expected threaded comment reply text"
    )
    assert any("Maximum number of ducks" in t for t in comment_texts), (
        "Expected threaded comment text"
    )

    for group in comment_groups:
        assert group.content_layer == ContentLayer.NOTES, (
            "Comments should be in NOTES content layer"
        )


def test_comment_cell_coordinates(documents) -> None:
    """Test that comment names include cell coordinates."""
    from docling_core.types.doc import GroupItem

    doc = next(item for path, item in documents if path.stem == "xlsx_comments")

    comment_groups = [
        g
        for g in doc.groups
        if isinstance(g, GroupItem) and g.name.startswith("comment-")
    ]

    # Should have 4 comments (2 notes + 2 threaded)
    assert len(comment_groups) == 4, (
        f"Expected 4 comment groups, got {len(comment_groups)}"
    )

    # Verify comment names include cell coordinates
    comment_names = [g.name for g in comment_groups]
    assert any("A1" in name for name in comment_names), "Expected comment for cell A1"
    assert any("B2" in name for name in comment_names), "Expected comment for cell B2"
    assert any("F7" in name for name in comment_names), (
        "Expected threaded comment for cell F7"
    )
    assert any("G12" in name for name in comment_names), (
        "Expected threaded comment for cell G12"
    )


def test_e2e_excel_conversions(documents, libreoffice_available) -> None:
    for gt_path, doc in documents:
        # xlsx_emf.xlsx contains EMF images that require LibreOffice to render.
        # Skip its groundtruth comparison when LibreOffice is not available.
        if gt_path.stem == "xlsx_emf" and not libreoffice_available:
            _log.info(
                "Skipping groundtruth comparison for %s: LibreOffice not available",
                gt_path.name,
            )
            continue

        included_content_layers = (
            set(ContentLayer) if gt_path.stem in "xlsx_comments" else None
        )
        my_layers = (
            included_content_layers
            if included_content_layers is not None
            else DEFAULT_CONTENT_LAYERS
        )
        pred_md: str = (
            MsExcelMarkdownDocSerializer(
                doc=doc,
                params=MarkdownParams(compact_tables=True, layers=my_layers),
            )
            .serialize()
            .text
        )
        assert verify_export(
            pred_md,
            str(gt_path) + ".md",
            GENERATE,
        ), "export to md"

        pred_itxt: str = doc._export_to_indented_text(
            max_text_len=70, explicit_tables=False
        )
        assert verify_export(pred_itxt, str(gt_path) + ".itxt", GENERATE), (
            "export to indented-text"
        )

        assert verify_document(doc, str(gt_path) + ".json", GENERATE), (
            "document document"
        )


def test_pages(documents) -> None:
    """Test the page count and page size of converted documents.

    Args:
        documents: The paths and converted documents.
    """
    # number of pages from the backend method
    path = next(item for item in get_excel_paths() if item.stem == "xlsx_01")
    in_doc = InputDocument(
        path_or_stream=path,
        format=InputFormat.XLSX,
        filename=path.stem,
        backend=MsExcelDocumentBackend,
    )
    backend = MsExcelDocumentBackend(in_doc=in_doc, path_or_stream=path)
    assert backend.page_count() == 4

    # number of pages from the converted document
    doc = next(item for path, item in documents if path.stem == "xlsx_01")
    assert len(doc.pages) == 4

    # page sizes as number of cells
    assert doc.pages.get(1).size.as_tuple() == (3.0, 7.0)
    assert doc.pages.get(2).size.as_tuple() == (16.0, 36.0)
    assert doc.pages.get(3).size.as_tuple() == (13.0, 36.0)
    # Sheet4 is hidden (ContentLayer.INVISIBLE) but still has real content
    assert doc.pages.get(4).size.as_tuple() == (1.0, 2.0)


def test_page_range() -> None:
    """Test that page_range selects a contiguous subset of sheets.

    xlsx_01.xlsx has 4 sheets. Converting with page_range=(2, 4) should yield
    only sheets 2-4, keeping their original page numbers (2, 3, 4).
    """
    path = next(item for item in get_excel_paths() if item.stem == "xlsx_01")

    converter = get_converter()
    doc = converter.convert(path, page_range=(2, 4)).document

    assert set(doc.pages.keys()) == {2, 3, 4}
    # original page numbering is preserved, so sizes match the full-document ones
    assert doc.pages.get(2).size.as_tuple() == (16.0, 36.0)
    assert doc.pages.get(3).size.as_tuple() == (13.0, 36.0)
    # Sheet4 is hidden (ContentLayer.INVISIBLE) but still has real content
    assert doc.pages.get(4).size.as_tuple() == (1.0, 2.0)


def test_page_range_with_sheet_names() -> None:
    """Test that page_range applies to the sheet_names-filtered set.

    With sheet_names dropping "Sheet2", the filtered sequence is
    [Sheet1, Sheet3, Sheet4] at positions 1, 2, 3. page_range=(2, 3) then
    selects Sheet3 and Sheet4 (pages 2 and 3 of the filtered set).
    """
    path = next(item for item in get_excel_paths() if item.stem == "xlsx_01")

    options = MsExcelBackendOptions(sheet_names=["Sheet1", "Sheet3", "Sheet4"])
    format_options = {InputFormat.XLSX: ExcelFormatOption(backend_options=options)}
    converter = DocumentConverter(
        allowed_formats=[InputFormat.XLSX], format_options=format_options
    )
    doc = converter.convert(path, page_range=(2, 3)).document

    assert set(doc.pages.keys()) == {2, 3}
    sheet_groups = [g.name for g in doc.groups if g.label == GroupLabel.SHEET]
    assert sheet_groups == ["Sheet3", "Sheet4"]


def test_chartsheet(documents) -> None:
    """Test that a native chart is parsed into a classified picture with data.

    ``parse_charts`` defaults to True, so the default converter extracts the
    "Duck Chart" bar chart. It should become a single PictureItem classified as a
    bar chart, captioned with the chart title, and carrying the chart's underlying
    data reconstructed as a table. The opt-out path is covered by
    ``test_chart_parsing_disabled``.

    Args:
        documents: The paths and converted documents.
    """
    doc = next(item for path, item in documents if path.stem == "xlsx_03_chartsheet")

    assert len(doc.pages) == 2
    assert doc.groups[1].name == "Duck Chart"

    # The chart anchors on the second sheet, so page 2 has a non-zero extent.
    assert doc.pages[2].size.width > 0
    assert doc.pages[2].size.height > 0

    pictures = list(doc.pictures)
    assert len(pictures) == 1, f"Expected one chart picture, got {len(pictures)}"

    picture = pictures[0]
    assert picture.prov[0].page_no == 2
    assert (
        picture.meta.classification.predictions[0].class_name
        == PictureClassificationLabel.BAR_CHART
    )
    assert picture.caption_text(doc) == "Wild Duck Observations by Year"

    # The two series and their shared categories are rebuilt as a table:
    #   | <blank> | Freshwater Ducks | Saltwater Ducks |
    #   | 2019    | 120              | 80              |
    #   ...
    #   | 2024    | 180              | 130             |
    chart_data = picture.meta.tabular_chart.chart_data
    assert (chart_data.num_rows, chart_data.num_cols) == (7, 3)
    grid = {
        (cell.start_row_offset_idx, cell.start_col_offset_idx): cell.text
        for cell in chart_data.table_cells
    }
    assert grid[(0, 1)] == "Freshwater Ducks"
    assert grid[(0, 2)] == "Saltwater Ducks"
    assert grid[(6, 0)] == "2024"
    assert grid[(6, 1)] == "180"
    assert grid[(6, 2)] == "130"


def test_chartsheet_data_values(documents) -> None:
    """Test that data values are extracted correctly from xlsx_03_chartsheet.

    This test verifies that calculated values (not formulas) are returned.
    The file contains duck observations with year 2024 having a total of 310 ducks.
    We need to verify that both 2024 and 310 appear in the parsed data.

    Args:
        documents: The paths and converted documents.
    """
    doc = next(item for path, item in documents if path.stem == "xlsx_03_chartsheet")

    # Find all tables
    tables = list(doc.tables)
    assert len(tables) > 0, "Should have at least one table"

    # Look for a table that has the year 2024 in it
    table_with_2024 = None
    row_index_of_2024 = None

    for table in tables:
        for cell in table.data.table_cells:
            if cell.text == "2024":
                table_with_2024 = table
                row_index_of_2024 = cell.start_row_offset_idx
                break
        if table_with_2024:
            break

    assert table_with_2024 is not None, "Should find a table containing year 2024"
    assert row_index_of_2024 is not None, "Should find row index for 2024"

    # Now verify that the value 310 exists in the document
    # (it may be in the same table or a different table due to how the parser splits tables)
    found_310 = False
    for table in tables:
        for cell in table.data.table_cells:
            if cell.text == "310":
                found_310 = True
                break
        if found_310:
            break

    assert found_310, "Should find the value 310 (total ducks for 2024) in the document"


def test_chart_parsing_disabled() -> None:
    """Test that parse_charts=False suppresses chart pictures.

    xlsx_03_chartsheet contains a single bar chart and no other images, so with
    chart parsing turned off the converted document has no pictures and the chart
    sheet's page keeps its empty extent.
    """
    path = next(item for item in get_excel_paths() if item.stem == "xlsx_03_chartsheet")

    options = MsExcelBackendOptions(parse_charts=False)
    format_options = {InputFormat.XLSX: ExcelFormatOption(backend_options=options)}
    converter = DocumentConverter(
        allowed_formats=[InputFormat.XLSX], format_options=format_options
    )
    doc = converter.convert(path).document

    assert len(list(doc.pictures)) == 0
    assert doc.pages[2].size.width == 0
    assert doc.pages[2].size.height == 0


def test_chart_image_rendering_disabled_by_default(documents) -> None:
    """Test that charts carry no rendered image unless the option is enabled.

    The default converter (used by the ``documents`` fixture) leaves
    render_chart_images=False, so the xlsx_03 chart picture keeps its
    classification and tabular data but no pixels. This guards the promise that
    the feature does not change default output for existing users.
    """
    doc = next(item for path, item in documents if path.stem == "xlsx_03_chartsheet")

    pictures = list(doc.pictures)
    assert len(pictures) == 1
    assert pictures[0].image is None, (
        "chart picture should have no image when render_chart_images is off"
    )


def test_chart_image_rendering(libreoffice_available) -> None:
    """Test that render_chart_images=True attaches a LibreOffice-rendered image.

    LibreOffice output is not byte-stable, and the cropped image size depends on
    the LibreOffice version and page setup, so the pixels are not compared
    against groundtruth. We assert the picture gains a non-trivial image while
    keeping the classification and tabular data extracted from the chart.

    Requires LibreOffice; skipped when it is not installed.
    """
    if not libreoffice_available:
        pytest.skip("LibreOffice is not installed — chart rendering cannot be tested")

    path = next(item for item in get_excel_paths() if item.stem == "xlsx_03_chartsheet")

    options = MsExcelBackendOptions(render_chart_images=True)
    format_options = {InputFormat.XLSX: ExcelFormatOption(backend_options=options)}
    converter = DocumentConverter(
        allowed_formats=[InputFormat.XLSX], format_options=format_options
    )
    doc = converter.convert(path).document

    pictures = list(doc.pictures)
    assert len(pictures) == 1, f"Expected one chart picture, got {len(pictures)}"

    picture = pictures[0]
    assert (
        picture.meta.classification.predictions[0].class_name
        == PictureClassificationLabel.BAR_CHART
    )
    assert picture.meta.tabular_chart is not None

    image = picture.get_image(doc=doc)
    assert image is not None, "chart picture should carry a rendered image"
    assert image.width > 50 and image.height > 50, (
        f"rendered chart image is implausibly small: {image.size}"
    )


def test_chart_render_does_not_mutate_source_chart() -> None:
    """Test that assembling the render workbook leaves the source chart intact.

    ``Worksheet.add_chart`` overwrites ``chart.anchor``. Were the backend to
    hand its own chart object to the temporary render workbook, the source
    chart's anchor would be replaced by a plain "A1" string and every later
    provenance bbox would silently collapse to (0, 0, 0, 0). Only the workbook
    assembly is exercised, so this runs without LibreOffice.
    """
    path = next(item for item in get_excel_paths() if item.stem == "xlsx_03_chartsheet")
    in_doc = InputDocument(
        path_or_stream=path,
        format=InputFormat.XLSX,
        filename=path.stem,
        backend=MsExcelDocumentBackend,
    )
    backend = MsExcelDocumentBackend(
        in_doc=in_doc,
        path_or_stream=path,
        options=MsExcelBackendOptions(render_chart_images=True),
    )
    chart = next(
        chart
        for name in backend.workbook.sheetnames
        for chart in backend.workbook[name]._charts
    )
    bbox_before = backend._anchor_to_tuple(chart.anchor)
    assert bbox_before != (0, 0, 0, 0), "test fixture should have a real anchor"

    assert backend._build_standalone_chart_workbook(chart) is not None

    assert backend._anchor_to_tuple(chart.anchor) == bbox_before, (
        "assembling the render workbook must not overwrite the source anchor"
    )


def test_inflated_rows_handling(documents) -> None:
    """Test that files with inflated max_row are handled correctly.

    xlsx_04_inflated.xlsx has inflated max_row (1,048,496) but only 7 rows of actual data.
    This test verifies that our backend correctly identifies true data bounds.
    """
    # First, verify the file has inflated max_row using openpyxl directly
    path = next(item for item in get_excel_paths() if item.stem == "xlsx_04_inflated")

    wb = load_workbook(path)
    ws = wb.active
    reported_max_row = ws.max_row

    # Assert that openpyxl reports inflated max_row
    assert reported_max_row > 100000, (
        f"xlsx_04_inflated.xlsx should have inflated max_row (expected >100k, got {reported_max_row:,}). "
        f"This test file is designed to verify proper handling of Excel files with inflated row counts."
    )

    _log.info(
        f"xlsx_04_inflated.xlsx - Openpyxl reported max_row: {reported_max_row:,}"
    )

    # Now test that our backend handles it correctly
    in_doc = InputDocument(
        path_or_stream=path,
        format=InputFormat.XLSX,
        filename=path.stem,
        backend=MsExcelDocumentBackend,
    )
    backend = MsExcelDocumentBackend(in_doc=in_doc, path_or_stream=path)

    # Verify backend detects correct number of pages (should be 4, like test-01)
    page_count = backend.page_count()
    assert page_count == 4, (
        f"Backend should detect 4 pages (same as test-01), got {page_count}"
    )

    # Verify converted document has correct pages
    doc = next(item for path, item in documents if path.stem == "xlsx_04_inflated")
    assert len(doc.pages) == 4, f"Document should have 4 pages, got {len(doc.pages)}"

    # Verify page sizes match expected dimensions (same as test-01)
    # These should reflect actual data, not inflated row counts
    assert doc.pages.get(1).size.as_tuple() == (3.0, 7.0), (
        f"Page 1 should be 3x7 cells, got {doc.pages.get(1).size.as_tuple()}"
    )
    assert doc.pages.get(2).size.as_tuple() == (16.0, 36.0), (
        f"Page 2 should be 16x36 cells, got {doc.pages.get(2).size.as_tuple()}"
    )
    assert doc.pages.get(3).size.as_tuple() == (13.0, 36.0), (
        f"Page 3 should be 13x36 cells, got {doc.pages.get(3).size.as_tuple()}"
    )
    # Sheet4 is hidden (ContentLayer.INVISIBLE) but still has real content
    assert doc.pages.get(4).size.as_tuple() == (1.0, 2.0), (
        f"Page 4 should be 1x2 cells (hidden sheet), got {doc.pages.get(4).size.as_tuple()}"
    )

    _log.info(
        f"✓ Successfully handled inflated max_row: "
        f"reported {reported_max_row:,} rows, "
        f"correctly processed as {page_count} pages with proper dimensions"
    )


def test_table_with_title():
    """Test that singleton cells with non-numeric content are treated as TextItem.

    When treat_singleton_as_text option is enabled, 1x1 tables containing non-numeric
    text should be converted to TextItem instead of TableItem. This test verifies that
    xlsx_05_table_with_title.xlsx is correctly parsed with this option.
    """
    path = next(
        item for item in get_excel_paths() if item.stem == "xlsx_05_table_with_title"
    )

    # Create converter with treat_singleton_as_text=True
    options = MsExcelBackendOptions(treat_singleton_as_text=True)
    format_options = {InputFormat.XLSX: ExcelFormatOption(backend_options=options)}
    converter = DocumentConverter(
        allowed_formats=[InputFormat.XLSX], format_options=format_options
    )

    conv_result: ConversionResult = converter.convert(path)
    doc: DoclingDocument = conv_result.document

    # With treat_singleton_as_text=True, the singleton title cell should be a TextItem
    texts = list(doc.texts)
    tables = list(doc.tables)

    assert len(texts) == 1, f"Should have 1 text item (the title), got {len(texts)}"
    assert len(tables) == 1, f"Should have 1 table, got {len(tables)}"

    # Verify the text item contains the title
    assert texts[0].text == "Number of freshwater ducks per year", (
        f"Text should be 'Number of freshwater ducks per year', got '{texts[0].text}'"
    )

    # Verify table dimensions
    table = tables[0]
    assert table.data.num_rows == 7, (
        f"Table should have 7 rows, got {table.data.num_rows}"
    )
    assert table.data.num_cols == 2, (
        f"Table should have 2 columns, got {table.data.num_cols}"
    )


def test_merged_section_label_above_table_preserves_column_headers() -> None:
    path = next(
        item
        for item in get_excel_paths()
        if item.stem == "xlsx_09_section_label_header"
    )
    headers = ["#", "Genre", "Sub-Genre", "Title", "Author", "Publisher", "Added"]

    converter = DocumentConverter(allowed_formats=[InputFormat.XLSX])
    doc = converter.convert(path).document

    assert [text.text for text in doc.texts] == ["Reading List"]
    assert len(doc.tables) == 1

    table = doc.tables[0]
    assert table.prov[0].bbox.t == 1
    assert table.data.num_rows == 3
    assert table.data.num_cols == len(headers)
    assert all(cell.text != "Reading List" for cell in table.data.table_cells)

    header_cells = [
        cell for cell in table.data.table_cells if cell.start_row_offset_idx == 0
    ]
    assert [cell.text for cell in header_cells] == headers
    assert all(cell.column_header for cell in header_cells)

    html = doc.export_to_html()
    assert '<th colspan="2">Reading List</th>' not in html
    assert "<th>#</th>" in html
    assert "<th>Genre</th>" in html


def test_split_leading_section_label_helper() -> None:
    backend = object.__new__(MsExcelDocumentBackend)

    no_split_table = ExcelTable(
        anchor=(2, 4),
        num_rows=1,
        num_cols=3,
        data=[
            ExcelCell(row=0, col=0, text="Reading List", row_span=1, col_span=2),
            ExcelCell(row=0, col=2, text="", row_span=1, col_span=1),
        ],
    )
    title_cell, unchanged_table = backend._split_leading_section_label(no_split_table)
    assert title_cell is None
    assert unchanged_table == no_split_table

    not_header_table = ExcelTable(
        anchor=(2, 4),
        num_rows=2,
        num_cols=3,
        data=[
            ExcelCell(row=0, col=0, text="Reading List", row_span=1, col_span=2),
            ExcelCell(row=0, col=1, text="", row_span=1, col_span=1),
            ExcelCell(row=0, col=2, text="", row_span=1, col_span=1),
            ExcelCell(row=1, col=0, text="Only one header", row_span=1, col_span=1),
            ExcelCell(row=1, col=1, text="", row_span=1, col_span=1),
            ExcelCell(row=1, col=2, text="", row_span=1, col_span=1),
        ],
    )
    title_cell, unchanged_table = backend._split_leading_section_label(not_header_table)
    assert title_cell is None
    assert unchanged_table == not_header_table

    split_table = ExcelTable(
        anchor=(2, 4),
        num_rows=3,
        num_cols=4,
        data=[
            ExcelCell(row=0, col=0, text="Reading List", row_span=1, col_span=2),
            ExcelCell(row=0, col=1, text="", row_span=1, col_span=1),
            ExcelCell(row=0, col=2, text="", row_span=1, col_span=1),
            ExcelCell(row=0, col=3, text="", row_span=1, col_span=1),
            ExcelCell(row=1, col=0, text="#", row_span=1, col_span=1),
            ExcelCell(row=1, col=1, text="Genre", row_span=1, col_span=1),
            ExcelCell(row=1, col=2, text="Sub-Genre", row_span=1, col_span=1),
            ExcelCell(row=1, col=3, text="Title", row_span=1, col_span=1),
            ExcelCell(row=2, col=0, text="1", row_span=1, col_span=1),
            ExcelCell(row=2, col=1, text="Fiction", row_span=1, col_span=1),
            ExcelCell(row=2, col=2, text="Mystery", row_span=1, col_span=1),
            ExcelCell(row=2, col=3, text="The Hound", row_span=1, col_span=1),
        ],
    )

    title_cell, split_result = backend._split_leading_section_label(split_table)

    assert title_cell is not None
    assert title_cell.text == "Reading List"
    assert split_result.anchor == (2, 5)
    assert split_result.num_rows == 2
    assert split_result.num_cols == 4
    assert [cell.row for cell in split_result.data] == [0, 0, 0, 0, 1, 1, 1, 1]
    assert [cell.text for cell in split_result.data[:4]] == [
        "#",
        "Genre",
        "Sub-Genre",
        "Title",
    ]


def test_bytesio_stream():
    """Test that Excel files can be loaded from BytesIO streams.

    This test verifies that the BytesIO code path in the backend is working correctly,
    ensuring that data_only=True is applied when loading workbooks from streams.
    """
    # Get a test Excel file
    path = next(item for item in get_excel_paths() if item.stem == "xlsx_01")

    # Load the file into a BytesIO stream
    buf = BytesIO(path.open("rb").read())

    # Create an InputDocument with the BytesIO stream
    in_doc = InputDocument(
        path_or_stream=buf,
        format=InputFormat.XLSX,
        filename=path.stem,
        backend=MsExcelDocumentBackend,
    )

    # Initialize the backend with the BytesIO stream
    backend = MsExcelDocumentBackend(in_doc=in_doc, path_or_stream=buf)

    # Verify the backend is valid
    assert backend.is_valid(), "Backend should be valid when loaded from BytesIO"

    # Verify page count matches expected value
    assert backend.page_count() == 4, "Should detect 4 pages from BytesIO stream"

    # Convert the document
    doc = backend.convert()

    # Verify the document was converted successfully
    assert doc is not None, "Document should be converted from BytesIO stream"
    assert len(doc.pages) == 4, "Document should have 4 pages"

    # Verify page sizes match expected dimensions
    assert doc.pages.get(1).size.as_tuple() == (3.0, 7.0)
    assert doc.pages.get(2).size.as_tuple() == (16.0, 36.0)
    assert doc.pages.get(3).size.as_tuple() == (13.0, 36.0)
    # Sheet4 is hidden (ContentLayer.INVISIBLE) but still has real content
    assert doc.pages.get(4).size.as_tuple() == (1.0, 2.0)


def test_edge_cases_merging() -> None:
    """Test that split tables are correctly merged using the region growing algorithm.

    Verifies:
    - Sheet 1 (missing_header): 1 table (Standard case)
    - Sheet 2 (Attached_left): 1 MERGED table (The critical fix!)
    - Sheet 3 (Diagonal): 2 separate tables (Correctly separated)
    """
    path = next(
        item for item in get_excel_paths() if item.stem == "xlsx_06_edge_cases_"
    )

    if not path.exists():
        pytest.skip(f"Test file {path} not found.")

    converter = DocumentConverter(allowed_formats=[InputFormat.XLSX])
    conv_result = converter.convert(path)
    doc = conv_result.document

    # Organize tables by Page Number (1-based index)
    tables_by_page = {}
    for table in doc.tables:
        p_no = table.prov[0].page_no
        if p_no not in tables_by_page:
            tables_by_page[p_no] = []
        tables_by_page[p_no].append(table)

    # Page 1: Standard table
    assert len(tables_by_page.get(1, [])) == 1, "Page 1 should have 1 table"

    # Page 2: The 'Attached left' case.
    # SUCCESS CONDITION: It is 1 single table.
    # (If the fix failed, this would be 2 tables).
    assert len(tables_by_page.get(2, [])) == 1, (
        f"Page 2 (Attached Left) should be 1 merged table, but found {len(tables_by_page.get(2, []))}"
    )

    # Page 3: Diagonal case.
    # These are physically separated by empty space, so they should remain 2 tables.
    assert len(tables_by_page.get(3, [])) == 2, (
        "Page 3 (Diagonal) should have 2 separate tables"
    )


def test_gap_tolerance_comparison() -> None:
    """Test the effect of gap_tolerance on table detection.

    Target: excel-tests.xlsx (Page 1), 'Power system' table.
    Structure: Col A ("1") | Col B (Empty) | Col C ("Rated system voltage")

    Verifies:
    1. Tolerance 0 (Default): The gap causes a split. The main data table starts at Col C.
    2. Tolerance 1: The gap is bridged. The table merges with Col A, starting at Col A.
    """
    path = next(
        item for item in get_excel_paths() if item.stem == "xlsx_07_gap_tolerance_"
    )
    if not path.exists():
        pytest.skip("Test file not found")

    # --- Helper to get the start column of the "Rated system voltage" table ---
    def get_table_start_col(tolerance: int) -> int:
        options = MsExcelBackendOptions(gap_tolerance=tolerance)
        format_options = {InputFormat.XLSX: ExcelFormatOption(backend_options=options)}

        converter = DocumentConverter(
            allowed_formats=[InputFormat.XLSX], format_options=format_options
        )
        doc = converter.convert(path).document
        print(doc)

        for table in doc.tables:
            # Check for unique text in the main body of the table
            texts = {cell.text for cell in table.data.table_cells}
            if "Rated system voltage" in texts:
                # Return the leftmost column index (0-based)
                return table.prov[0].bbox.l

        pytest.fail(f"Could not find 'Power system' table with tolerance={tolerance}")

    # --- ASSERTION 1: Strict Behavior (gap_tolerance=0) ---
    # The empty Col B should split the table.
    # The text "Rated system voltage" is in Col C (Index 2).
    start_col_strict = get_table_start_col(0)
    assert start_col_strict == 2, (
        f"Default (0) tolerance should split the table. "
        f"Expected start at Col C (2), got {start_col_strict}"
    )

    # --- ASSERTION 2: Merged Behavior (gap_tolerance=1) ---
    # The empty Col B should be ignored.
    # The table should merge left to include "1" in Col A (Index 0).
    start_col_merged = get_table_start_col(1)
    assert start_col_merged == 0, (
        f"Tolerance 1 should merge the table. "
        f"Expected start at Col A (0), got {start_col_merged}"
    )


def test_one_cell_anchor_image():
    """Test that images with OneCellAnchor are positioned correctly.

    OneCellAnchor images (the default when inserting images in Excel) should
    use the anchor cell as the bounding box origin, not default to (0,0,0,0).
    """
    path = next(
        item for item in get_excel_paths() if item.stem == "xlsx_08_one_cell_anchor"
    )

    converter = get_converter()
    conv_result = converter.convert(path)
    doc = conv_result.document

    pictures = list(doc.pictures)
    assert len(pictures) == 1, f"Should have 1 picture, got {len(pictures)}"

    prov = pictures[0].prov[0]
    # Image was placed at cell D2 (col=3, row=1 in 0-based)
    assert prov.bbox.l == 3.0, f"Image left should be 3.0 (col D), got {prov.bbox.l}"
    assert prov.bbox.t == 1.0, f"Image top should be 1.0 (row 2), got {prov.bbox.t}"
    assert prov.bbox.r == 4.0, f"Image right should be 4.0, got {prov.bbox.r}"
    assert prov.bbox.b == 2.0, f"Image bottom should be 2.0, got {prov.bbox.b}"


def test_find_data_tables_handles_a_filled_last_excel_row(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1048576"] = "last row"
    file_path = tmp_path / "test.xlsx"
    workbook.save(file_path)

    in_doc = InputDocument(
        path_or_stream=file_path,
        format=InputFormat.XLSX,
        filename=file_path.stem,
        backend=MsExcelDocumentBackend,
    )
    backend = MsExcelDocumentBackend(in_doc=in_doc, path_or_stream=file_path)
    doc: DoclingDocument = backend.convert()

    tables = doc.tables
    assert len(tables) == 1

    table = tables[0]
    print(table)
    assert table.prov[0].bbox.t == 1048575
    assert table.data.num_rows == 1
    assert table.data.num_cols == 1
    assert len(table.data.table_cells) == 1
    assert table.data.table_cells[0].text == "last row"


def test_emf_images_in_xlsx(libreoffice_available):
    """Test that EMF images embedded in XLSX files are extracted via LibreOffice.

    The test file xlsx_emf.xlsx contains three sheets:
      - 'Raster in emf'  - a raster image stored as EMF (openpyxl drops these)
      - 'Vector in emf'  - a vector image stored as EMF (openpyxl drops these)
      - 'Raster in webp' - a regular PNG image (openpyxl handles these normally)

    On every sheet the image sits above a small table (image at rows 1-10,
    table at rows 11-13), so the picture must appear before the table in the
    exported document.

    Requires LibreOffice for the EMF sheets; skipped when it is not installed.
    """
    if not libreoffice_available:
        pytest.skip("LibreOffice is not installed — EMF conversion cannot be tested")

    path = next(item for item in get_excel_paths() if item.stem == "xlsx_emf")

    converter = get_converter()
    conv_result = converter.convert(path)
    doc = conv_result.document

    # Three sheets → three pages, each with one picture and one table
    assert doc.num_pages() == 3

    pictures = list(doc.pictures)
    tables = list(doc.tables)
    assert len(pictures) == 3, (
        f"Expected 3 pictures (one per sheet), got {len(pictures)}"
    )
    assert len(tables) == 3, f"Expected 3 tables (one per sheet), got {len(tables)}"

    # All pictures must carry image data (i.e. not be empty placeholders)
    for pic in pictures:
        assert pic.image is not None, (
            f"Picture on page {pic.prov[0].page_no} has no image data"
        )

    # On every page the picture must come before the table in document order
    items_by_page: dict[int, list] = {}
    for item, _ in doc.iterate_items(traverse_pictures=True):
        if not item.prov:
            continue
        page_no = item.prov[0].page_no
        items_by_page.setdefault(page_no, []).append(item)

    for page_no, items in items_by_page.items():
        pic_indices = [i for i, it in enumerate(items) if isinstance(it, PictureItem)]
        tbl_indices = [i for i, it in enumerate(items) if isinstance(it, TableItem)]
        assert pic_indices and tbl_indices, f"Page {page_no} missing picture or table"
        assert max(pic_indices) < min(tbl_indices), (
            f"Page {page_no}: picture (idx {pic_indices}) should come before "
            f"table (idx {tbl_indices}) in document order"
        )
