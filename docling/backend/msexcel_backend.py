from __future__ import annotations

import collections
import logging
import posixpath
import shutil
import subprocess
import warnings
from copy import deepcopy
from datetime import datetime
from io import BytesIO
from pathlib import Path
from tempfile import mkdtemp
from typing import Annotated, Any, Callable, Final, cast
from zipfile import ZipFile

import pypdfium2
from docling_core.types.doc import (
    BoundingBox,
    ContentLayer,
    CoordOrigin,
    DocItem,
    DocItemLabel,
    DoclingDocument,
    DocumentOrigin,
    GroupItem,
    GroupLabel,
    ImageRef,
    PictureClassificationLabel,
    PictureClassificationMetaField,
    PictureClassificationPrediction,
    PictureMeta,
    ProvenanceItem,
    Size,
    TableCell,
    TableData,
    TabularChartMetaField,
)
from lxml import etree
from PIL import Image as PILImage, UnidentifiedImageError
from pydantic import BaseModel, Field, NonNegativeInt, PositiveInt
from pydantic.dataclasses import dataclass
from typing_extensions import override

from docling.backend.abstract_backend import (
    DeclarativeDocumentBackend,
    PaginatedDocumentBackend,
)
from docling.backend.docx.drawingml.utils import (
    crop_whitespace,
    get_libreoffice_cmd,
)
from docling.datamodel.backend_options import MsExcelBackendOptions
from docling.datamodel.base_models import InputFormat
from docling.datamodel.document import InputDocument
from docling.exceptions import DocumentLoadError

_log = logging.getLogger(__name__)

_OPENPYXL_AVAILABLE: bool = False
_OPENPYXL_IMPORT_ERROR: ImportError | None = None
try:  # pragma: no cover - import-time guard
    from openpyxl import Workbook, load_workbook
    from openpyxl.chartsheet.chartsheet import Chartsheet
    from openpyxl.drawing.image import Image
    from openpyxl.drawing.spreadsheet_drawing import (
        OneCellAnchor,
        SpreadsheetDrawing,
        TwoCellAnchor,
    )
    from openpyxl.packaging.relationship import get_dependents, get_rels_path
    from openpyxl.styles import PatternFill
    from openpyxl.utils.cell import range_boundaries
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.xml.constants import IMAGE_NS

    _OPENPYXL_AVAILABLE = True
except ImportError as e:  # pragma: no cover - import-time guard
    _OPENPYXL_IMPORT_ERROR = e

_INSTALL_HINT = (
    "The 'openpyxl' package is required to process Excel files. "
    "Install it with `pip install 'docling-slim[format-xlsx]'`."
)


# Safe XML parser — prevents XXE, DTD-over-network, and entity-expansion attacks.
_SAFE_XML_PARSER: Final = etree.XMLParser(
    resolve_entities=False,
    load_dtd=False,
    no_network=True,
    dtd_validation=False,
)

_CHART_RENDER_HINT = (
    "LibreOffice is required to render Excel charts as images "
    "(render_chart_images=True). Install LibreOffice and make sure `soffice` is "
    "on PATH. Charts still keep their classification and reconstructed tabular "
    "data without it."
)

# Maps an openpyxl chart object's ``tagname`` (the DrawingML element name, e.g.
# "barChart") to the docling picture-classification label we tag the emitted
# PictureItem with. Chart types not listed fall back to OTHER_CHART.
_CHART_TAGNAME_TO_CLASSIFICATION: Final[dict[str, PictureClassificationLabel]] = {
    "barChart": PictureClassificationLabel.BAR_CHART,
    "bar3DChart": PictureClassificationLabel.BAR_CHART,
    "lineChart": PictureClassificationLabel.LINE_CHART,
    "line3DChart": PictureClassificationLabel.LINE_CHART,
    "pieChart": PictureClassificationLabel.PIE_CHART,
    "pie3DChart": PictureClassificationLabel.PIE_CHART,
    "doughnutChart": PictureClassificationLabel.PIE_CHART,
    "scatterChart": PictureClassificationLabel.SCATTER_CHART,
    "areaChart": PictureClassificationLabel.OTHER_CHART,
    "area3DChart": PictureClassificationLabel.OTHER_CHART,
}


def _has_unsafe_zip_paths(namelist: list[str]) -> bool:
    """Return True if any ZIP member name is absolute or contains a path traversal."""
    return any(m.startswith("/") or ".." in m for m in namelist)


@dataclass
class DataRegion:
    """Represents the bounding rectangle of non-empty cells in a worksheet."""

    min_row: Annotated[
        PositiveInt, Field(description="Smallest row index (1-based index).")
    ]
    max_row: Annotated[
        PositiveInt, Field(description="Largest row index (1-based index).")
    ]
    min_col: Annotated[
        PositiveInt, Field(description="Smallest column index (1-based index).")
    ]
    max_col: Annotated[
        PositiveInt, Field(description="Largest column index (1-based index).")
    ]

    def width(self) -> PositiveInt:
        """Number of columns in the data region."""
        return self.max_col - self.min_col + 1

    def height(self) -> PositiveInt:
        """Number of rows in the data region."""
        return self.max_row - self.min_row + 1


class ExcelCell(BaseModel):
    """Represents an Excel cell.

    Attributes:
        row: The row number of the cell.
        col: The column number of the cell.
        text: The text content of the cell.
        row_span: The number of rows the cell spans.
        col_span: The number of columns the cell spans.
    """

    row: int
    col: int
    text: str
    row_span: int
    col_span: int


class ExcelTable(BaseModel):
    """Represents an Excel table on a worksheet.

    Attributes:
        anchor: The column and row indices of the upper-left cell of the table
        (0-based index).
        num_rows: The number of rows in the table.
        num_cols: The number of columns in the table.
        data: The data in the table, represented as a list of ExcelCell objects.
    """

    anchor: tuple[NonNegativeInt, NonNegativeInt]
    num_rows: int
    num_cols: int
    data: list[ExcelCell]


class MsExcelDocumentBackend(DeclarativeDocumentBackend, PaginatedDocumentBackend):
    """Backend for parsing Excel workbooks.

    The backend converts an Excel workbook into a DoclingDocument object.
    Each worksheet is converted into a separate page.
    The following elements are parsed:
    - Cell contents, parsed as tables. If two groups of cells are disconnected
      between each other, they will be parsed as two different tables.
    - Images, parsed as PictureItem objects.
    - Cell comments (notes), both old-style and Excel 365+ threaded comments.

    The DoclingDocument tables and pictures have their provenance information, including
    the position in their original Excel worksheet. The position is represented by a
    bounding box object with the cell indices as units (0-based index). The size of this
    bounding box is the number of columns and rows that the table or picture spans.

    Limitations:
        - Threaded comments (Excel 365+) are only extracted when the file is provided
          as a Path. When provided as a BytesIO stream, threaded comments cannot be
          extracted because the stream is consumed by openpyxl during initialization.
          Old-style cell comments (notes) are always extracted regardless of input type.
    """

    # Maximum seconds to wait for a single LibreOffice EMF/WMF conversion.
    # Raise this value if conversions time out on unusually large or complex files.
    LIBREOFFICE_TIMEOUT_S: Final[int] = 60

    # Maximum uncompressed byte sizes accepted when reading members from the XLSX zip.
    # These caps guard against decompression-bomb payloads in drawing XML / image files.
    _MAX_DRAWING_BYTES: Final[int] = 10 * 1024 * 1024  # 10 MB
    _MAX_IMAGE_BYTES: Final[int] = 50 * 1024 * 1024  # 50 MB

    @override
    def __init__(
        self,
        in_doc: InputDocument,
        path_or_stream: BytesIO | Path,
        options: MsExcelBackendOptions | None = None,
    ) -> None:
        """Initialize the MsExcelDocumentBackend object.

        Parameters:
            in_doc: The input document object.
            path_or_stream: The path or stream to the Excel file.
            options: Backend options for Excel parsing.

        Raises:
            RuntimeError: An error occurred parsing the file.
        """
        if not _OPENPYXL_AVAILABLE:
            raise ImportError(_INSTALL_HINT) from _OPENPYXL_IMPORT_ERROR
        if options is None:
            options = MsExcelBackendOptions()
        super().__init__(in_doc, path_or_stream, options)

        self.page_range = in_doc.limits.page_range

        # Current sheet group; set at the start of each sheet conversion
        self.parent: GroupItem | None = None

        # Lazy-initialized LibreOffice converter for EMF/WMF images
        self.xlsx_to_pdf_converter: Callable | None = None
        self.xlsx_to_pdf_converter_init: bool = False

        self.workbook = None
        try:
            # Suppress the openpyxl warning for WMF/EMF images being dropped:
            # those formats are handled separately via LibreOffice conversion.
            with warnings.catch_warnings():
                warnings.filterwarnings(
                    "ignore",
                    message=r".* image format is not supported so the image is being dropped",
                    category=UserWarning,
                    module=r"openpyxl\.reader\.drawings",
                )
                warnings.filterwarnings(
                    "ignore",
                    message=r"The image .* will be removed because it cannot be read",
                    category=UserWarning,
                    module=r"openpyxl\.reader\.drawings",
                )
                if isinstance(self.path_or_stream, BytesIO):
                    self.workbook = load_workbook(
                        filename=self.path_or_stream, data_only=True
                    )
                elif isinstance(self.path_or_stream, Path):
                    self.workbook = load_workbook(
                        filename=str(self.path_or_stream), data_only=True
                    )

            self.valid = self.workbook is not None
        except Exception as e:
            self.valid = False

            raise DocumentLoadError(
                f"MsExcelDocumentBackend could not load document with hash {self.document_hash}"
            ) from e

    def _parse_threaded_comments(
        self, sheet_name: str
    ) -> dict[str, tuple[str, str, datetime | None]]:
        """Parse threaded comments from Excel XML for a specific sheet.

        Returns a dict mapping cell coordinates to (author, text, timestamp) tuples.
        Only works when path_or_stream is a Path (not BytesIO).

        Security Note:
            Uses secure XML parser configuration to prevent XXE attacks and validates
            ZIP file paths to prevent zip-slip attacks.
        """
        threaded_comments: dict[str, tuple[str, str, datetime | None]] = {}

        # Only extract from Path objects (BytesIO is consumed by load_workbook)
        if not isinstance(self.path_or_stream, Path):
            return threaded_comments

        # Namespace for threaded comments XML
        ns = {
            "tc": "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"
        }

        try:
            if isinstance(self.path_or_stream, BytesIO):
                self.path_or_stream.seek(0)

            with ZipFile(self.path_or_stream, "r") as zip_file:
                if _has_unsafe_zip_paths(zip_file.namelist()):
                    _log.warning("Skipping file with unsafe ZIP paths")
                    return threaded_comments

                person_map: dict[str, str] = {}
                try:
                    person_xml = zip_file.read("xl/persons/person.xml")
                    person_tree = etree.fromstring(person_xml, parser=_SAFE_XML_PARSER)
                    person_map = {
                        person.get("id"): person.get("displayName")
                        for person in person_tree.findall(".//tc:person", namespaces=ns)
                        if person.get("id") and person.get("displayName")
                    }
                except Exception as e:
                    _log.debug(f"Could not parse person.xml: {e}")

                sheet_num = next(
                    (
                        i
                        for i, ws in enumerate(self.workbook.worksheets, 1)
                        if ws.title == sheet_name
                    ),
                    None,
                )
                if sheet_num is None:
                    return threaded_comments

                threaded_file = f"xl/threadedComments/threadedComment{sheet_num}.xml"
                try:
                    threaded_xml = zip_file.read(threaded_file)
                    threaded_tree = etree.fromstring(
                        threaded_xml, parser=_SAFE_XML_PARSER
                    )

                    for comment in threaded_tree.findall(
                        ".//tc:threadedComment", namespaces=ns
                    ):
                        cell_ref = comment.get("ref")
                        text_elem = comment.find("tc:text", namespaces=ns)

                        if cell_ref and text_elem is not None:
                            text = text_elem.text or ""
                            author = person_map.get(comment.get("personId"), "Unknown")

                            timestamp = None
                            if timestamp_str := comment.get("dT"):
                                try:
                                    # Normalize timestamp for Python 3.10 compatibility
                                    # xlsx uses fractional seconds with variable precision
                                    normalized = timestamp_str.replace("Z", "+00:00")
                                    if "." in normalized and "+" in normalized:
                                        parts = normalized.split(".")
                                        frac_and_tz = parts[1].split("+")
                                        frac = frac_and_tz[0].ljust(6, "0")[:6]
                                        normalized = (
                                            f"{parts[0]}.{frac}+{frac_and_tz[1]}"
                                        )
                                    elif "." in normalized:
                                        parts = normalized.split(".")
                                        frac = parts[1].ljust(6, "0")[:6]
                                        normalized = f"{parts[0]}.{frac}"
                                    timestamp = datetime.fromisoformat(normalized)
                                except Exception as e:
                                    _log.debug(
                                        f"Could not parse timestamp '{timestamp_str}': {e}"
                                    )

                            threaded_comments[cell_ref] = (author, text, timestamp)

                except Exception as e:
                    _log.debug(f"Could not parse {threaded_file}: {e}")

        except Exception as e:
            _log.debug(f"Could not parse threaded comments: {e}")

        return threaded_comments

    @override
    def is_valid(self) -> bool:
        _log.debug(f"valid: {self.valid}")
        return self.valid

    @classmethod
    @override
    def supports_pagination(cls) -> bool:
        return True

    @override
    def page_count(self) -> int:
        if self.is_valid() and self.workbook:
            sheet_names_filter: list[str] | None = (
                self.options.sheet_names
                if isinstance(self.options, MsExcelBackendOptions)
                else None
            )
            if sheet_names_filter is None:
                return len(self.workbook.sheetnames)
            return sum(
                1 for name in self.workbook.sheetnames if name in sheet_names_filter
            )
        else:
            return 0

    @classmethod
    @override
    def supported_formats(cls) -> set[InputFormat]:
        return {InputFormat.XLSX}

    @override
    def convert(self) -> DoclingDocument:
        """Parse the Excel workbook into a DoclingDocument object.

        Raises:
            RuntimeError: Unable to run the conversion since the backend object failed to
            initialize.

        Returns:
            The DoclingDocument object representing the Excel workbook.
        """
        origin = DocumentOrigin(
            filename=self.file.name or "file.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            binary_hash=self.document_hash,
        )

        doc = DoclingDocument(name=self.file.stem or "file.xlsx", origin=origin)

        if self.is_valid():
            doc = self._convert_workbook(doc)
        else:
            raise RuntimeError(
                f"Cannot convert doc with {self.document_hash} because the backend failed to init."
            )

        return doc

    def _convert_workbook(self, doc: DoclingDocument) -> DoclingDocument:
        """Parse the Excel workbook and attach its structure to a DoclingDocument.

        Args:
            doc: A DoclingDocument object.

        Returns:
            A DoclingDocument object with the parsed items.
        """

        if self.workbook is not None:
            sheet_names_filter: list[str] | None = (
                self.options.sheet_names
                if isinstance(self.options, MsExcelBackendOptions)
                else None
            )

            start_page, end_page = self.page_range

            page_no = 0
            # Iterate over all sheets
            for idx, name in enumerate(self.workbook.sheetnames):
                if sheet_names_filter is not None and name not in sheet_names_filter:
                    _log.debug(f"Skipping sheet {idx}: {name} (filtered out)")
                    continue

                # Page numbers are 1-based positions within the (optionally filtered)
                # set of sheets. Increment before the range check so selected sheets
                # keep their original page numbers (e.g. page_range=(2, 4) -> 2, 3, 4).
                page_no += 1
                if page_no < start_page or page_no > end_page:
                    _log.debug(
                        f"Skipping sheet {idx}: {name} "
                        f"(page {page_no} outside range {start_page}-{end_page})"
                    )
                    continue

                _log.info(f"Processing sheet {idx}: {name} as page {page_no}")

                sheet = self.workbook[name]
                # do not rely on sheet.max_column, sheet.max_row if there are images
                page = doc.add_page(page_no=page_no, size=Size(width=0, height=0))

                self.parent = doc.add_group(
                    parent=None,
                    label=GroupLabel.SHEET,
                    name=name,
                    content_layer=self._get_sheet_content_layer(sheet),
                )
                doc = self._convert_sheet(doc, sheet, page_no)
                width, height = self._find_page_size(doc, page_no)
                page.size = Size(width=width, height=height)

            if sheet_names_filter is not None:
                unmatched = set(sheet_names_filter) - set(self.workbook.sheetnames)
                if unmatched:
                    _log.warning(
                        "sheet_names filter contains names not found in workbook: %s",
                        sorted(unmatched),
                    )
        else:
            _log.error("Workbook is not initialized.")

        return doc

    def _convert_sheet(
        self, doc: DoclingDocument, sheet: Worksheet | Chartsheet, page_no: int
    ) -> DoclingDocument:
        """Parse an Excel worksheet and attach its structure to a DoclingDocument

        Args:
            doc: The DoclingDocument to be updated.
            sheet: The Excel worksheet to be parsed.
            page_no: The dense (1-based) page number for this sheet in the output document.

        Returns:
            The updated DoclingDocument.
        """
        if isinstance(sheet, Worksheet):
            doc = self._find_tables_in_sheet(doc, sheet, page_no)
            doc = self._find_images_in_sheet(doc, sheet, page_no)
        # Charts can be on both Worksheet and Chartsheet objects
        if isinstance(sheet, (Worksheet, Chartsheet)):
            doc = self._find_chart_in_sheet(doc, sheet, page_no)
        self._sort_sheet_children_by_position(doc, page_no)

        return doc

    def _sort_sheet_children_by_position(
        self, doc: DoclingDocument, page_no: int
    ) -> None:
        """Sort the current sheet group's direct children by top-row position.

        Tables are added before images during sheet conversion.  When an image
        sits above a table on the sheet (smaller row index), it would otherwise
        appear after the table in the exported document.  Sorting the sheet
        group's children by their ``bbox.t`` corrects the visual order.

        Children without provenance on the current page sort to the end.

        Args:
            doc: The DoclingDocument whose current sheet group is sorted in place.
            page_no: The 1-based page number of the sheet being processed.
        """
        sheet_group = self.parent
        if sheet_group is None:
            return

        def _top_row(ref: Any) -> float:
            item = ref.resolve(doc)
            if item is None:
                return float("inf")
            for prov in getattr(item, "prov", []):
                if prov.page_no == page_no:
                    return prov.bbox.t
            return float("inf")

        sheet_group.children.sort(key=_top_row)

    def _find_tables_in_sheet(
        self, doc: DoclingDocument, sheet: Worksheet, page_no: int
    ) -> DoclingDocument:
        """Find all tables in an Excel sheet and attach them to a DoclingDocument.

        Also extracts comments from cells and links them to their corresponding table cells.

        Args:
            doc: The DoclingDocument to be updated.
            sheet: The Excel worksheet to be parsed.
            page_no: The dense (1-based) page number for this sheet in the output document.

        Returns:
            The updated DoclingDocument.
        """

        if self.workbook is not None:
            content_layer = self._get_sheet_content_layer(sheet)
            tables, comment_map = self._find_data_tables(sheet)

            treat_singleton_as_text = (
                isinstance(self.options, MsExcelBackendOptions)
                and self.options.treat_singleton_as_text
            )

            for excel_table in tables:
                title_cell, excel_table = self._split_leading_section_label(excel_table)
                origin_col = excel_table.anchor[0]
                origin_row = excel_table.anchor[1]
                num_rows = excel_table.num_rows
                num_cols = excel_table.num_cols
                if title_cell is not None:
                    doc.add_text(
                        text=title_cell.text,
                        label=DocItemLabel.TEXT,
                        parent=self.parent,
                        prov=ProvenanceItem(
                            page_no=page_no,
                            charspan=(0, 0),
                            bbox=BoundingBox.from_tuple(
                                (
                                    origin_col + title_cell.col,
                                    origin_row - 1,
                                    origin_col + title_cell.col + title_cell.col_span,
                                    origin_row,
                                ),
                                origin=CoordOrigin.TOPLEFT,
                            ),
                        ),
                        content_layer=content_layer,
                    )

                if treat_singleton_as_text and len(excel_table.data) == 1:
                    doc.add_text(
                        text=excel_table.data[0].text,
                        label=DocItemLabel.TEXT,
                        parent=self.parent,
                        prov=ProvenanceItem(
                            page_no=page_no,
                            charspan=(0, 0),
                            bbox=BoundingBox.from_tuple(
                                (
                                    origin_col,
                                    origin_row,
                                    origin_col + num_cols,
                                    origin_row + num_rows,
                                ),
                                origin=CoordOrigin.TOPLEFT,
                            ),
                        ),
                        content_layer=content_layer,
                    )
                else:
                    table_data = TableData(
                        num_rows=num_rows,
                        num_cols=num_cols,
                        table_cells=[],
                    )

                    for excel_cell in excel_table.data:
                        cell = TableCell(
                            text=excel_cell.text,
                            row_span=excel_cell.row_span,
                            col_span=excel_cell.col_span,
                            start_row_offset_idx=excel_cell.row,
                            end_row_offset_idx=excel_cell.row + excel_cell.row_span,
                            start_col_offset_idx=excel_cell.col,
                            end_col_offset_idx=excel_cell.col + excel_cell.col_span,
                            column_header=excel_cell.row == 0,
                            row_header=False,
                        )
                        table_data.table_cells.append(cell)

                    doc.add_table(
                        data=table_data,
                        parent=self.parent,
                        prov=ProvenanceItem(
                            page_no=page_no,
                            charspan=(0, 0),
                            bbox=BoundingBox.from_tuple(
                                (
                                    origin_col,
                                    origin_row,
                                    origin_col + num_cols,
                                    origin_row + num_rows,
                                ),
                                origin=CoordOrigin.TOPLEFT,
                            ),
                        ),
                        content_layer=content_layer,
                    )

            # Extract comments and link them to cells
            for (row, col), (author, raw_text, timestamp) in comment_map.items():
                metadata_parts = []
                if author:
                    metadata_parts.append(f"author: {author}")
                if timestamp:
                    timestamp_str = timestamp.isoformat(timespec="milliseconds")
                    metadata_parts.append(f"time: {timestamp_str}")

                if metadata_parts and raw_text:
                    full_text = f"[{', '.join(metadata_parts)}]: {raw_text}"
                elif metadata_parts:
                    full_text = f"[{', '.join(metadata_parts)}]"
                else:
                    full_text = raw_text

                cell_item = self._find_cell_item(doc, page_no, row, col)
                targets = [cell_item] if cell_item else None

                comment_group = doc.add_group(
                    label=GroupLabel.COMMENT_SECTION,
                    name=f"comment-{sheet.title}-{sheet.cell(row=row + 1, column=col + 1).coordinate}",
                    content_layer=ContentLayer.NOTES,
                )
                doc.add_comment(
                    text=full_text,
                    targets=targets,
                    parent=comment_group,
                )

                if not targets:
                    _log.debug(
                        f"Comment at {sheet.title}!{sheet.cell(row=row + 1, column=col + 1).coordinate} "
                        f"has no cell item for linking"
                    )

        return doc

    def _split_leading_section_label(
        self, table: ExcelTable
    ) -> tuple[ExcelCell | None, ExcelTable]:
        """Split a merged section label from an adjacent data table."""
        if table.num_rows < 2 or table.num_cols < 2:
            return None, table

        first_row_cells = [cell for cell in table.data if cell.row == 0]
        first_row_text_cells = [cell for cell in first_row_cells if cell.text.strip()]
        if len(first_row_text_cells) != 1:
            return None, table

        title_cell = first_row_text_cells[0]
        if (
            title_cell.col != 0
            or title_cell.row_span != 1
            or title_cell.col_span <= 1
            or title_cell.col_span > table.num_cols
        ):
            return None, table

        second_row_header_cells = [
            cell
            for cell in table.data
            if cell.row == 1 and cell.text.strip() and cell.col_span == 1
        ]
        if len(second_row_header_cells) < 2:
            return None, table

        data = [
            ExcelCell(
                row=cell.row - 1,
                col=cell.col,
                text=cell.text,
                row_span=cell.row_span,
                col_span=cell.col_span,
            )
            for cell in table.data
            if cell.row > 0
        ]
        return (
            title_cell,
            ExcelTable(
                anchor=(table.anchor[0], table.anchor[1] + 1),
                num_rows=table.num_rows - 1,
                num_cols=table.num_cols,
                data=data,
            ),
        )

    def _find_true_data_bounds(self, sheet: Worksheet) -> DataRegion:
        """Find the true data boundaries (min/max rows and columns) in a worksheet.

        This function scans all cells to find the smallest rectangular region that contains
        all non-empty cells or merged cell ranges. It returns the minimal and maximal
        row/column indices that bound the actual data region.

        Args:
            sheet: The worksheet to analyze.

        Returns:
            A data region representing the smallest rectangle that covers all data and merged cells.
            If the sheet is empty, returns (1, 1, 1, 1) by default.
        """
        min_row, min_col = None, None
        max_row, max_col = 0, 0

        for cell in sheet._cells.values():
            if cell.value is not None:
                r, c = cell.row, cell.column
                min_row = r if min_row is None else min(min_row, r)
                min_col = c if min_col is None else min(min_col, c)
                max_row = max(max_row, r)
                max_col = max(max_col, c)

        # Expand bounds to include merged cells
        for merged in sheet.merged_cells.ranges:
            min_row = (
                merged.min_row if min_row is None else min(min_row, merged.min_row)
            )
            min_col = (
                merged.min_col if min_col is None else min(min_col, merged.min_col)
            )
            max_row = max(max_row, merged.max_row)
            max_col = max(max_col, merged.max_col)

        # If no data found, default to (1, 1, 1, 1)
        if min_row is None or min_col is None:
            min_row = min_col = max_row = max_col = 1

        return DataRegion(min_row, max_row, min_col, max_col)

    def _find_data_tables(
        self, sheet: Worksheet
    ) -> tuple[
        list[ExcelTable], dict[tuple[int, int], tuple[str, str, datetime | None]]
    ]:
        """Find all compact rectangular data tables in an Excel worksheet.

        Also collects comments from cells during the same iteration for efficiency.

        Args:
            sheet: The Excel worksheet to be parsed.

        Returns:
            A tuple containing:
                - A list of ExcelTable objects representing the data tables
                - A dict mapping (row, col) to (author, comment_text, timestamp) for cells with comments
        """
        bounds: DataRegion = self._find_true_data_bounds(
            sheet
        )  # The true data boundaries
        tables: list[ExcelTable] = []  # List to store found tables
        visited: set[tuple[int, int]] = set()  # Track already visited cells
        comment_map: dict[
            tuple[int, int], tuple[str, str, datetime | None]
        ] = {}  # Collect comments

        # Parse threaded comments from XML (Excel 365+ format with proper author names and timestamps)
        threaded_comments = self._parse_threaded_comments(sheet.title)

        # Limit scan to actual data bounds
        for ri, row in enumerate(
            sheet.iter_rows(
                min_row=bounds.min_row,
                max_row=bounds.max_row,
                min_col=bounds.min_col,
                max_col=bounds.max_col,
                values_only=False,
            ),
            start=bounds.min_row - 1,
        ):
            for rj, cell in enumerate(row, start=bounds.min_col - 1):
                # Collect comment if present
                if cell.comment is not None:
                    author = cell.comment.author or ""
                    raw_text = (
                        str(cell.comment.text).strip() if cell.comment.text else ""
                    )
                    timestamp = None

                    # Check if this is a threaded comment with better data in XML
                    cell_coord = cell.coordinate
                    if cell_coord in threaded_comments:
                        author, raw_text, timestamp = threaded_comments[cell_coord]
                    elif author.startswith("tc={") and "[Threaded comment]" in raw_text:
                        # Fallback: extract from openpyxl's text if XML parsing failed
                        if "Comment:\n" in raw_text:
                            raw_text = raw_text.split("Comment:\n", 1)[1].strip()
                        author = "Threaded comment"

                    if raw_text:
                        comment_map[(ri, rj)] = (author, raw_text, timestamp)

                if cell.value is None or (ri, rj) in visited:
                    continue

                # If the cell starts a new table, find its bounds
                table_bounds, visited_cells = self._find_table_bounds(
                    sheet, ri, rj, bounds.max_row, bounds.max_col
                )
                visited.update(visited_cells)  # Mark these cells as visited
                tables.append(table_bounds)

        return tables, comment_map

    def _find_table_bounds(
        self,
        sheet: Worksheet,
        start_row: int,
        start_col: int,
        max_row: int,
        max_col: int,
    ) -> tuple[ExcelTable, set[tuple[int, int]]]:
        """Determine table bounds using a Flood Fill (BFS) strategy.

        This method identifies contiguous regions of non-empty cells in an Excel worksheet
        using a breadth-first search algorithm. It accurately detects non-rectangular tables
        (e.g., L-shapes, staggered columns) by exploring connected cells with optional gap
        tolerance.

        The algorithm operates in two phases:
        1. Flood Fill: Uses BFS to find all connected cells starting from the given position
        2. Data Extraction: Builds a rectangular bounding box and extracts cell data,
           handling merged cells appropriately

        Args:
            sheet: The Excel worksheet to analyze.
            start_row: The starting row index (0-based) for the flood fill.
            start_col: The starting column index (0-based) for the flood fill.
            max_row: The exclusive row bound to consider in the worksheet.
            max_col: The exclusive column bound to consider in the worksheet.

        Returns:
            A tuple containing:
                - ExcelTable: An object representing the detected table with its anchor
                  position, dimensions, and cell data.
                - set[tuple[int, int]]: A set of (row, col) tuples representing all cells
                  that were visited during the flood fill, used to prevent re-scanning.

        Note:
            The method respects the GAP_TOLERANCE option, which allows cells separated by
            empty cells to be considered part of the same table if within tolerance distance.
        """
        GAP_TOLERANCE = cast(MsExcelBackendOptions, self.options).gap_tolerance

        # Queue for BFS: (row, col)
        queue = collections.deque([(start_row, start_col)])

        # Track local visited for this specific table to avoid loops
        # (The caller maintains a global 'visited' set to avoid re-starting tables)
        table_cells: set[tuple[int, int]] = set()
        table_cells.add((start_row, start_col))

        # Track bounds dynamically
        min_r, max_r = start_row, start_row
        min_c, max_c = start_col, start_col

        # Helper: Check if a cell has content
        def has_content(r, c):
            if r < 0 or c < 0 or r >= max_row or c >= max_col:
                return False

            # 1. Check direct value
            cell = sheet.cell(row=r + 1, column=c + 1)
            if cell.value is not None:
                return True

            # 2. Check merge ranges
            for mr in sheet.merged_cells.ranges:
                if cell.coordinate in mr:
                    return True
            return False

        # --- Phase 1: Flood Fill (Connectivity Check) ---
        while queue:
            curr_r, curr_c = queue.popleft()

            # Update Bounds
            min_r = min(min_r, curr_r)
            max_r = max(max_r, curr_r)
            min_c = min(min_c, curr_c)
            max_c = max(max_c, curr_c)

            # Check neighbors in 4 directions
            # We respect GAP_TOLERANCE by looking 'k' steps away
            directions = [
                (0, 1),
                (0, -1),
                (1, 0),
                (-1, 0),
            ]

            for dr, dc in directions:
                # Check neighbors up to tolerance distance
                # We prioritize the closest neighbor.
                for step in range(1, GAP_TOLERANCE + 2):
                    nr, nc = curr_r + (dr * step), curr_c + (dc * step)

                    if (nr, nc) in table_cells:
                        break  # Already part of this table, don't jump over it

                    if has_content(nr, nc):
                        table_cells.add((nr, nc))
                        queue.append((nr, nc))
                        # Found a connection in this direction, stop extending 'gap'
                        break

        # --- Phase 2: Extract Data (Semantic Grid) ---
        data = []

        # We must identify cells that are "shadowed" by a merge (not the top-left)
        hidden_merge_cells = set()
        for mr in sheet.merged_cells.ranges:
            mr_min_r, mr_min_c = mr.min_row - 1, mr.min_col - 1
            mr_max_r, mr_max_c = mr.max_row - 1, mr.max_col - 1
            for r in range(mr_min_r, mr_max_r + 1):
                for c in range(mr_min_c, mr_max_c + 1):
                    if r == mr_min_r and c == mr_min_c:
                        continue
                    hidden_merge_cells.add((r, c))

        # We iterate the bounding box of the found region
        # Gaps inside the bounding box become empty cells (preserving layout)
        for ri in range(min_r, max_r + 1):
            for rj in range(min_c, max_c + 1):
                # If this cell was part of the flood-fill OR is inside the bounds
                # (We include gaps inside the bounds to keep the table rectangular)

                # Logic: If we found a "U" shape, do we fill the middle?
                # Yes, Excel tables are typically treated as rectangular bounding boxes.

                if (ri, rj) in hidden_merge_cells:
                    continue

                cell = sheet.cell(row=ri + 1, column=rj + 1)
                cell_text = str(cell.value) if cell.value is not None else ""

                # Compute Spans
                row_span = 1
                col_span = 1
                for mr in sheet.merged_cells.ranges:
                    if (ri + 1) == mr.min_row and (rj + 1) == mr.min_col:
                        row_span = (mr.max_row - mr.min_row) + 1
                        col_span = (mr.max_col - mr.min_col) + 1
                        break

                data.append(
                    ExcelCell(
                        row=ri - min_r,
                        col=rj - min_c,
                        text=cell_text,
                        row_span=row_span,
                        col_span=col_span,
                    )
                )

        # The 'visited_cells' returned to the caller MUST strictly be the ones
        # that contain data/merges, so the main loop doesn't re-scan them.
        # However, to avoid overlapping tables, we should mark the whole bbox?
        # Standard behavior: Mark the specific connected cells we found.
        return (
            ExcelTable(
                anchor=(min_c, min_r),
                num_rows=max_r + 1 - min_r,
                num_cols=max_c + 1 - min_c,
                data=data,
            ),
            table_cells,
        )

    @staticmethod
    def _anchor_to_tuple(anchor: Any) -> tuple[int, int, int, int]:
        """Convert an openpyxl anchor object to a (left_col, top_row, right_col, bottom_row) tuple.

        Args:
            anchor: A TwoCellAnchor, OneCellAnchor, or unknown anchor type.

        Returns:
            A 4-tuple suitable for BoundingBox.from_tuple.
        """
        if isinstance(anchor, TwoCellAnchor):
            return (
                anchor._from.col,
                anchor._from.row,
                anchor.to.col + 1,
                anchor.to.row + 1,
            )
        if isinstance(anchor, OneCellAnchor):
            return (
                anchor._from.col,
                anchor._from.row,
                anchor._from.col + 1,
                anchor._from.row + 1,
            )
        return (0, 0, 0, 0)

    def _get_libreoffice_converter(self) -> Callable | None:
        """Lazily initialize and return a LibreOffice converter callable.

        The converter accepts ``(input_path: Path, output_path: Path)`` and
        converts the input file to PDF at the given output path.

        Returns:
            A converter callable, or None when LibreOffice is not available.
        """
        if self.xlsx_to_pdf_converter_init:
            return self.xlsx_to_pdf_converter

        self.xlsx_to_pdf_converter_init = True
        libreoffice_cmd = get_libreoffice_cmd()
        if libreoffice_cmd is None:
            _log.debug(
                "LibreOffice not found — EMF/WMF images in XLSX will be skipped."
            )
            self.xlsx_to_pdf_converter = None
            return None

        def _convert(input_path: Path, output_path: Path) -> None:
            subprocess.run(
                [
                    libreoffice_cmd,
                    "--headless",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    str(output_path.parent),
                    str(input_path),
                ],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=True,
                timeout=self.LIBREOFFICE_TIMEOUT_S,
            )
            # LibreOffice names the output after the input stem
            expected = output_path.parent / (input_path.stem + ".pdf")
            if expected != output_path:
                expected.rename(output_path)

        self.xlsx_to_pdf_converter = _convert
        return self.xlsx_to_pdf_converter

    def _convert_emf_to_pil(self, image_bytes: bytes) -> PILImage.Image | None:
        """Convert a raw EMF or WMF image to a PIL Image via LibreOffice.

        LibreOffice can convert standalone ``.emf``/``.wmf`` files to PDF
        directly — no DOCX or XLSX wrapper is needed.  The raw bytes are
        written to a temp file, converted to PDF, and the first page is
        rendered with pypdfium2.

        Args:
            image_bytes: Raw EMF or WMF image data.

        Returns:
            A PIL Image, or None if LibreOffice is unavailable or conversion fails.
        """
        converter = self._get_libreoffice_converter()
        if converter is None:
            return None

        # WMF placeable magic: D7 CD C6 9A; everything else we treat as EMF.
        suffix = ".wmf" if image_bytes[:4] == b"\xd7\xcd\xc6\x9a" else ".emf"
        temp_dir = Path(mkdtemp())
        try:
            input_path = temp_dir / f"image{suffix}"
            output_path = temp_dir / "image.pdf"
            input_path.write_bytes(image_bytes)
            converter(input_path, output_path)
            if not output_path.exists():
                _log.debug("LibreOffice produced no PDF output for %s", input_path.name)
                return None
            pdf = pypdfium2.PdfDocument(str(output_path))
            page = pdf[0]
            pil_image = crop_whitespace(page.render(scale=2).to_pil())
            page.close()
            pdf.close()
            return pil_image
        except Exception as exc:
            _log.debug("EMF/WMF conversion via LibreOffice failed: %s", exc)
            return None
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def _find_unsupported_images_in_sheet(
        self, doc: DoclingDocument, sheet: Worksheet, page_no: int
    ) -> DoclingDocument:
        """Find EMF/WMF images dropped by openpyxl and add them via LibreOffice.

        openpyxl silently drops WMF images and raises ``OSError`` for EMF
        images (which PIL cannot decode natively).  This method re-parses the
        drawing relationships for the sheet directly from the XLSX zip archive
        and converts those unsupported images using LibreOffice.

        Args:
            doc: The DoclingDocument to be updated.
            sheet: The Excel worksheet to be parsed.
            page_no: The dense (1-based) page number for this sheet in the output document.

        Returns:
            The updated DoclingDocument.
        """
        # drawing rels are stored on the worksheet object by openpyxl
        drawing_paths: list[str] = [
            rel.target
            for rel in sheet._rels.find(SpreadsheetDrawing._rel_type)  # type: ignore[attr-defined]
        ]
        if not drawing_paths:
            return doc

        content_layer = self._get_sheet_content_layer(sheet)

        if isinstance(self.path_or_stream, BytesIO):
            self.path_or_stream.seek(0)

        try:
            with ZipFile(self.path_or_stream, "r") as zf:
                if _has_unsafe_zip_paths(zf.namelist()):
                    _log.warning(
                        "Skipping EMF/WMF scan: XLSX archive contains unsafe ZIP paths"
                    )
                    return doc

                for drawing_path in drawing_paths:
                    doc = self._process_drawing_for_unsupported_images(
                        doc, zf, drawing_path, page_no, content_layer
                    )
        except Exception as exc:
            _log.debug("Could not scan drawing files for unsupported images: %s", exc)

        return doc

    def _process_drawing_for_unsupported_images(
        self,
        doc: DoclingDocument,
        zf: ZipFile,
        drawing_path: str,
        page_no: int,
        content_layer: ContentLayer | None,
    ) -> DoclingDocument:
        """Scan one drawing XML file and convert any EMF/WMF blips found.

        Args:
            doc: The DoclingDocument to update.
            zf: Open ZipFile for the xlsx archive.
            drawing_path: Absolute path inside the zip to the drawing XML.
            page_no: Page number (1-based) for provenance.
            content_layer: ContentLayer for added pictures.

        Returns:
            The updated DoclingDocument.
        """
        if drawing_path not in zf.namelist():
            return doc

        tree = etree.fromstring(zf.read(drawing_path), parser=_SAFE_XML_PARSER)
        try:
            drawing = SpreadsheetDrawing.from_tree(tree)
        except TypeError:
            return doc

        rels_path = get_rels_path(drawing_path)
        if rels_path not in zf.namelist():
            return doc
        deps = get_dependents(zf, rels_path)

        for rel in drawing._blip_rels:
            dep = deps.get(rel.embed)
            if dep.Type != IMAGE_NS or dep.target not in zf.namelist():
                continue

            image_bytes = zf.read(dep.target)

            # Skip images PIL can already handle — openpyxl already added those.
            try:
                pil_probe = PILImage.open(BytesIO(image_bytes))
                probe_buf = BytesIO()
                pil_probe.save(probe_buf, format="PNG")
                continue  # PIL succeeded; openpyxl took care of this one
            except (UnidentifiedImageError, OSError):
                pass

            pil_image = self._convert_emf_to_pil(image_bytes)
            if pil_image is None:
                _log.warning(
                    "Could not convert unsupported image '%s'. "
                    "Install LibreOffice for EMF/WMF support in XLSX files.",
                    dep.target,
                )
                continue

            doc.add_picture(
                parent=self.parent,
                image=ImageRef.from_pil(image=pil_image, dpi=72),
                caption=None,
                prov=ProvenanceItem(
                    page_no=page_no,
                    charspan=(0, 0),
                    bbox=BoundingBox.from_tuple(
                        self._anchor_to_tuple(rel.anchor),
                        origin=CoordOrigin.TOPLEFT,
                    ),
                ),
                content_layer=content_layer,
            )

        return doc

    def _find_images_in_sheet(
        self, doc: DoclingDocument, sheet: Worksheet, page_no: int
    ) -> DoclingDocument:
        """Find images in the Excel sheet and attach them to the DoclingDocument.

        Args:
            doc: The DoclingDocument to be updated.
            sheet: The Excel worksheet to be parsed.
            page_no: The dense (1-based) page number for this sheet in the output document.

        Returns:
            The updated DoclingDocument.
        """
        if self.workbook is not None:
            content_layer = self._get_sheet_content_layer(sheet)
            # Images that PIL can read are already loaded by openpyxl into sheet._images
            for item in sheet._images:  # type: ignore[attr-defined]
                try:
                    image: Image = cast(Image, item)
                    ref = image.ref
                    pil_image = (
                        ref if isinstance(ref, PILImage.Image) else PILImage.open(ref)
                    )
                    doc.add_picture(
                        parent=self.parent,
                        image=ImageRef.from_pil(image=pil_image, dpi=72),
                        caption=None,
                        prov=ProvenanceItem(
                            page_no=page_no,
                            charspan=(0, 0),
                            bbox=BoundingBox.from_tuple(
                                self._anchor_to_tuple(image.anchor),
                                origin=CoordOrigin.TOPLEFT,
                            ),
                        ),
                        content_layer=content_layer,
                    )
                except Exception:
                    _log.error("could not extract the image from excel sheets")

            # EMF/WMF images are silently dropped by openpyxl; handle them separately
            doc = self._find_unsupported_images_in_sheet(doc, sheet, page_no)

        return doc

    def _find_chart_in_sheet(
        self, doc: DoclingDocument, sheet: Worksheet | Chartsheet, page_no: int
    ) -> DoclingDocument:
        """Find native charts on a sheet and attach them as classified pictures.

        openpyxl parses each embedded chart (bar, line, pie, scatter, ...) into
        ``sheet._charts``.  For every chart we emit a PictureItem whose meta
        carries (a) the chart-type classification and (b) the chart's underlying
        numbers reconstructed as a TableData.  This mirrors the ODF backend so
        XLSX and ODS charts have the same downstream shape.

        Args:
            doc: The DoclingDocument to update.
            sheet: The worksheet or chart sheet being parsed.
            page_no: The 1-based page number of this sheet, used for provenance.

        Returns:
            The updated DoclingDocument.
        """

        if not (
            isinstance(self.options, MsExcelBackendOptions)
            and self.options.parse_charts
        ):
            return doc
        content_layer = self._get_sheet_content_layer(sheet)

        charts = sheet._charts  # type: ignore[attr-defined]
        if not charts:
            return doc

        # Rendering a chart to an actual image is opt-in and needs LibreOffice.
        render_charts = self.options.render_chart_images
        if render_charts and self._get_libreoffice_converter() is None:
            _log.warning(_CHART_RENDER_HINT)
            render_charts = False

        for chart in charts:
            try:
                classification = _CHART_TAGNAME_TO_CLASSIFICATION.get(
                    chart.tagname, PictureClassificationLabel.OTHER_CHART
                )
                caption_text = self._chart_title_text(chart)
                table_data = self._chart_to_table_data(chart)

                bbox = BoundingBox.from_tuple(
                    self._anchor_to_tuple(chart.anchor),
                    origin=CoordOrigin.TOPLEFT,
                )

                # On any rendering failure fall back to the no-image behavior:
                # the picture still carries its classification and chart data.
                image_ref = None
                if render_charts:
                    try:
                        chart_image = self._render_chart_image(chart)
                        if chart_image is not None:
                            image_ref = ImageRef.from_pil(image=chart_image, dpi=72)
                    except Exception:
                        _log.warning(
                            "could not render a chart image; keeping chart data "
                            "without image",
                            exc_info=True,
                        )

                caption_item = (
                    doc.add_text(
                        label=DocItemLabel.CAPTION,
                        text=caption_text,
                        content_layer=content_layer,
                    )
                    if caption_text
                    else None
                )

                picture = doc.add_picture(
                    parent=self.parent,
                    image=image_ref,
                    caption=caption_item,
                    prov=ProvenanceItem(
                        page_no=page_no,
                        charspan=(0, 0),
                        bbox=bbox,
                    ),
                    content_layer=content_layer,
                )

                picture.meta = PictureMeta(
                    classification=PictureClassificationMetaField(
                        predictions=[
                            PictureClassificationPrediction(class_name=classification)
                        ]
                    ),
                    tabular_chart=(
                        TabularChartMetaField(chart_data=table_data)
                        if table_data is not None
                        else None
                    ),
                )
            except Exception:
                _log.error(
                    "could not extract a chart from the excel sheet", exc_info=True
                )

        return doc

    @staticmethod
    def _chart_title_text(chart: Any) -> str | None:
        """Extract the plain-text title of an openpyxl chart, if any.

        A chart title is stored as DrawingML rich text: ``chart.title`` is a
        Title object whose ``.tx.rich.p`` is a list of paragraphs, each with a
        list of runs ``.r``, each run carrying its text in ``.t``.  We flatten
        all runs into a single string.  Returns None when the chart has no title.

        Args:
            chart: An openpyxl chart object (BarChart, LineChart, ...).

        Returns:
            The concatenated title text, or None.
        """
        title = chart.title
        if title is None:
            return None

        if isinstance(title, str):
            return title or None
        tx = title.tx
        if tx is None or tx.rich is None:
            return None

        runs: list[str] = []
        for paragraph in tx.rich.p:
            for run in paragraph.r or []:
                if run.t:
                    runs.append(run.t)
        text = "".join(runs).strip()
        return text or None

    def _chart_to_table_data(self, chart: Any) -> TableData | None:
        """Reconstruct a chart's underlying data grid as a TableData.

        Layout produced (categories down the first column, one column per series):

            | <blank> | <series 0 name> | <series 1 name> | ...
            | cat_0   | val_0,0         | val_1,0         | ...
            | cat_1   | val_0,1         | val_1,1         | ...

        Chart data is stored in openpyxl as *references* back into the workbook
        (e.g. "'Sheet1'!$B$2:$B$7"), which we resolve to cached cell values.
        Scatter charts use ``xVal``/``yVal`` instead of ``cat``/``val``.

        Args:
            chart: An openpyxl chart object.

        Returns:
            A TableData, or None if the chart exposes no usable series.
        """
        series_list = list(chart.series)
        if not series_list:
            return None

        categories: list[str] = []
        for series in series_list:
            cat_ref = self._ref_formula(series.cat) or self._ref_formula(series.xVal)
            if cat_ref:
                categories = self._resolve_reference(cat_ref)
                break

        columns: list[tuple[str, list[str]]] = []

        for series in series_list:
            value_ref = self._ref_formula(series.val) or self._ref_formula(series.yVal)
            values = self._resolve_reference(value_ref) if value_ref else []

            name_ref = self._ref_formula(series.tx)
            if name_ref:
                resolved = self._resolve_reference(name_ref)
                name = resolved[0] if resolved else ""
            elif series.tx is not None and series.tx.v is not None:
                name = str(series.tx.v)
            else:
                name = ""
            columns.append((name, values))

        num_data_rows = max([len(categories)] + [len(values) for _, values in columns])
        if num_data_rows == 0:
            return None

        num_rows = num_data_rows + 1
        num_cols = 1 + len(columns)
        cells: list[TableCell] = []

        header_labels = [""] + [name for name, _ in columns]
        for col_idx, label in enumerate(header_labels):
            cells.append(
                TableCell(
                    text=label,
                    row_span=1,
                    col_span=1,
                    start_row_offset_idx=0,
                    end_row_offset_idx=1,
                    start_col_offset_idx=col_idx,
                    end_col_offset_idx=col_idx + 1,
                    column_header=True,
                    row_header=False,
                )
            )
        for data_row in range(num_data_rows):
            row_idx = data_row + 1
            category = categories[data_row] if data_row < len(categories) else ""
            row_texts = [category] + [
                (values[data_row] if data_row < len(values) else "")
                for _, values in columns
            ]
            for col_idx, text in enumerate(row_texts):
                cells.append(
                    TableCell(
                        text=text,
                        row_span=1,
                        col_span=1,
                        start_row_offset_idx=row_idx,
                        end_row_offset_idx=row_idx + 1,
                        start_col_offset_idx=col_idx,
                        end_col_offset_idx=col_idx + 1,
                        column_header=False,
                        row_header=(col_idx == 0),
                    )
                )

        return TableData(num_rows=num_rows, num_cols=num_cols, table_cells=cells)

    @staticmethod
    def _to_float(text: str) -> float | None:
        """Parse a cell string into a float, or None if it is not numeric.

        Cached values arrive as strings; a chart plots them only when they are
        written back as numbers. Non-numeric cells (blank, labels) return None
        so the caller can leave them as text.
        """
        try:
            return float(text)
        except (TypeError, ValueError):
            return None

    def _render_chart_image(self, chart: Any) -> PILImage.Image | None:
        """Render a native chart to an image via LibreOffice.

        XLSX stores charts as vector definitions with no embedded raster.  To
        obtain a picture we isolate the chart into a throwaway single-chart
        workbook (its data copied onto hidden sheets so the series still
        resolve), convert that to PDF with LibreOffice — the same external tool
        already used for EMF/WMF images — and rasterize the first page with
        pypdfium2, trimming the surrounding whitespace.

        Args:
            chart: An openpyxl chart object.

        Returns:
            A PIL Image, or None when LibreOffice is unavailable, the chart has
            no resolvable data, or the conversion fails.
        """
        converter = self._get_libreoffice_converter()
        if converter is None:
            return None

        standalone = self._build_standalone_chart_workbook(chart)
        if standalone is None:
            return None

        temp_dir = Path(mkdtemp())
        try:
            input_path = temp_dir / "chart.xlsx"
            output_path = temp_dir / "chart.pdf"
            standalone.save(input_path)
            converter(input_path, output_path)
            if not output_path.exists():
                _log.debug("LibreOffice produced no PDF output for a chart")
                return None
            pdf = pypdfium2.PdfDocument(str(output_path))
            page = pdf[0]
            pil_image = crop_whitespace(page.render(scale=2).to_pil())
            page.close()
            pdf.close()
            return pil_image
        except Exception as exc:
            _log.debug("Chart rendering via LibreOffice failed: %s", exc)
            return None
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def _build_standalone_chart_workbook(self, chart: Any) -> Workbook | None:
        """Build a one-chart workbook with the chart's data on hidden sheets.

        A chart references its data by sheet-qualified ranges (e.g.
        ``"'Sheet1'!$B$2:$B$7"``).  Every referenced range is recreated here —
        same sheet name, same cell coordinates, cached values — on hidden
        sheets, and the chart is anchored alone on a single visible sheet.
        LibreOffice prints only the visible sheet, so the resulting PDF holds
        just the chart.

        Args:
            chart: An openpyxl chart object.

        Returns:
            A populated Workbook, or None when the chart exposes no resolvable
            data references.
        """
        if self.workbook is None:
            return None

        references: list[str] = []
        for series in chart.series:
            for data_source in (
                series.cat,
                series.xVal,
                series.val,
                series.yVal,
                series.tx,
            ):
                formula = self._ref_formula(data_source)
                if formula:
                    references.append(formula)
        if not references:
            return None

        standalone = Workbook()
        # A fresh Workbook always holds exactly one worksheet; take it directly
        # rather than via ``.active``, which is typed as optional.
        chart_sheet = standalone.worksheets[0]
        chart_sheet.title = "chart_render"

        # Every reference must be copied, so this cannot short-circuit.
        copied = [self._copy_reference_into(standalone, ref) for ref in references]
        if not any(copied):
            return None

        # ``add_chart`` overwrites ``chart.anchor``; copy so the workbook's own
        # chart keeps the anchor its provenance bbox is derived from.
        chart_sheet.add_chart(deepcopy(chart), "A1")
        return standalone

    def _copy_reference_into(self, target: Workbook, ref: str) -> bool:
        """Copy one chart data range into ``target`` on a hidden sheet.

        Recreates the referenced sheet by name if needed, hides it, and writes
        the cached cell values at their original coordinates.  Numeric-looking
        text is coerced to numbers so the chart plots it as values rather than
        labels.

        Args:
            target: The standalone Workbook being assembled.
            ref: A sheet-qualified range reference, e.g. "'Sheet1'!$B$2:$B$7".

        Returns:
            True when the referenced range was resolved and copied.
        """
        if self.workbook is None or "!" not in ref:
            return False

        sheet_part, cell_range = ref.rsplit("!", 1)
        sheet_part = sheet_part.strip()
        if sheet_part.startswith("'") and sheet_part.endswith("'"):
            sheet_part = sheet_part[1:-1].replace("''", "'")
        if sheet_part not in self.workbook.sheetnames:
            _log.debug("Chart references unknown sheet %r", sheet_part)
            return False

        try:
            bounds = range_boundaries(cell_range)
        except Exception:
            _log.debug("Could not parse chart range %r", cell_range)
            return False
        if any(bound is None for bound in bounds):
            # Open-ended ranges (e.g. "B:B") carry no usable row bounds.
            _log.debug("Chart range %r is not fully bounded", cell_range)
            return False
        min_col, min_row, max_col, max_row = cast(tuple[int, int, int, int], bounds)

        source_sheet = self.workbook[sheet_part]
        if sheet_part in target.sheetnames:
            dest_sheet = target[sheet_part]
        else:
            dest_sheet = target.create_sheet(sheet_part)
            dest_sheet.sheet_state = "hidden"

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                value = source_sheet.cell(row=row, column=col).value
                if isinstance(value, str):
                    numeric = self._to_float(value)
                    if numeric is not None:
                        value = numeric
                dest_sheet.cell(row=row, column=col, value=value)
        return True

    @staticmethod
    def _ref_formula(data_source: Any) -> str | None:
        """Return the cell-range formula string from a chart data source.

        A chart's series title/categories/values are each a small openpyxl
        object (SeriesLabel, AxDataSource, NumDataSource) that may hold either a
        numeric reference (``.numRef``) or a string reference (``.strRef``); both
        expose the range formula on ``.f`` (e.g. "'Sheet1'!$B$2:$B$7").  These
        objects don't share a common base exposing both attributes, so we probe
        each — a narrowly-scoped getattr against a third-party API.

        Args:
            data_source: A chart data-source object, or None.

        Returns:
            The range-formula string, or None if absent.
        """
        if data_source is None:
            return None
        num_ref = getattr(data_source, "numRef", None)
        if num_ref is not None and num_ref.f:
            return num_ref.f
        str_ref = getattr(data_source, "strRef", None)
        if str_ref is not None and str_ref.f:
            return str_ref.f
        return None

    def _resolve_reference(self, ref: str) -> list[str]:
        """Resolve a chart range reference to a flat list of cell-value strings.

        Charts point at their data by reference rather than embedding it, e.g.
        ``"'Duck Observations'!$B$2:$B$7"``.  We split off the (possibly quoted)
        sheet name, convert the range to bounds with openpyxl's
        ``range_boundaries``, and read the *cached* values from the workbook
        (the backend loads with ``data_only=True``, so we get computed numbers,
        not formulas).  Values are returned in row-major order.

        Args:
            ref: A range reference string, optionally sheet-qualified.

        Returns:
            The referenced cell values as strings ("" for empty cells).  Returns
            an empty list if the sheet is missing (e.g. filtered out) or the
            reference can't be parsed.
        """
        if self.workbook is None:
            return []

        if "!" in ref:
            sheet_part, cell_range = ref.rsplit("!", 1)
            sheet_part = sheet_part.strip()
            if sheet_part.startswith("'") and sheet_part.endswith("'"):
                sheet_part = sheet_part[1:-1].replace("''", "'")
            sheet_name = sheet_part
        else:
            sheet_name, cell_range = self.workbook.active.title, ref
        if sheet_name not in self.workbook.sheetnames:
            _log.debug("Chart references unknown sheet %r", sheet_name)
            return []

        target_sheet = self.workbook[sheet_name]

        try:
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        except Exception:
            _log.debug("Could not parse chart range %r", cell_range)
            return []

        values: list[str] = []
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                value = target_sheet.cell(row=row, column=col).value
                values.append("" if value is None else str(value))

        return values

    @staticmethod
    def _find_page_size(
        doc: DoclingDocument, page_no: PositiveInt
    ) -> tuple[float, float]:
        """Return (width, height) for the given page in cell-index units.

        Width and height are the maximum ``r`` and ``b`` bbox coordinates seen
        across all items on the page, regardless of content layer.  Because
        bboxes use ``CoordOrigin.TOPLEFT`` with the sheet origin at ``(0, 0)``,
        the page extent equals the largest right/bottom value — not the span
        between the leftmost/topmost and rightmost/bottommost edges.
        """
        width: float = 0.0
        height: float = 0.0
        for item, _ in doc.iterate_items(
            traverse_pictures=True,
            page_no=page_no,
            included_content_layers=set(ContentLayer),
        ):
            if not isinstance(item, DocItem):
                continue
            for provenance in item.prov:
                if provenance.page_no != page_no:
                    continue
                width = max(width, provenance.bbox.r)
                height = max(height, provenance.bbox.b)

        return (width, height)

    def _find_cell_item(
        self, doc: DoclingDocument, page_no: int, row: int, col: int
    ) -> DocItem | None:
        """Find the DocItem (table cell or text) at the given row/col position.

        Args:
            doc: The DoclingDocument to search.
            page_no: The page number to search in.
            row: Row index (0-based).
            col: Column index (0-based).

        Returns:
            The DocItem at that position, or None if not found.
        """
        for item, _ in doc.iterate_items(page_no=page_no):
            if not isinstance(item, DocItem):
                continue

            for prov in item.prov:
                if prov.page_no != page_no:
                    continue

                bbox = prov.bbox
                # Check if cell position is within this item's bounding box
                if bbox.l <= col < bbox.r and bbox.t <= row < bbox.b:
                    return item

        return None

    @staticmethod
    def _get_sheet_content_layer(sheet: Worksheet) -> ContentLayer | None:
        return (
            None
            if sheet.sheet_state == Worksheet.SHEETSTATE_VISIBLE
            else ContentLayer.INVISIBLE
        )
