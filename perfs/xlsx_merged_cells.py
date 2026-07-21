"""Benchmark full XLSX conversion for worksheets with many merged ranges.

The script generates temporary workbooks containing horizontal ``A:C`` merges and
an ordinary value in column ``D``, then converts each workbook through
``DocumentConverter``. Each result is emitted as one JSON object per line and
includes generation time, conversion time, and peak Python memory usage.

Run from the repository root, for example::

    uv run python perfs/xlsx_merged_cells.py --merge-count 100 1000
    uv run python perfs/xlsx_merged_cells.py --merge-count 5000 --output results.jsonl

Progress is written to stderr so stdout remains valid JSONL.
"""

from __future__ import annotations

import argparse
import json
import platform
import re
import subprocess
import sys
import time
import tracemalloc
from dataclasses import asdict, dataclass
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook
from tqdm import tqdm

from docling.datamodel.base_models import InputFormat
from docling.document_converter import DocumentConverter

DEFAULT_MERGE_COUNTS = (100, 1_000, 5_000, 20_000)


@dataclass(frozen=True)
class RuntimeInfo:
    python_version: str
    platform: str
    docling_version: str
    commit_sha: str


@dataclass(frozen=True)
class BenchmarkResult:
    python_version: str
    platform: str
    docling_version: str
    commit_sha: str
    merge_count: int
    non_empty_cells: int
    rectangle_area: int
    generation_seconds: float
    conversion_seconds: float
    peak_python_memory_mb: float
    table_count: int


def _positive_int(value: str) -> int:
    parsed = int(value)
    if parsed <= 0:
        raise argparse.ArgumentTypeError("merge counts must be positive integers")
    return parsed


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--merge-count",
        type=_positive_int,
        nargs="+",
        default=DEFAULT_MERGE_COUNTS,
        help="Merged-range counts to benchmark.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Optional JSONL output path; results are always written to stdout.",
    )
    return parser.parse_args()


def _runtime_info() -> RuntimeInfo:
    repository_root = Path(__file__).resolve().parents[1]
    pyproject_text = (repository_root / "pyproject.toml").read_text(encoding="utf-8")
    version_match = re.search(
        r'^version\s*=\s*"([^"]+)"', pyproject_text, flags=re.MULTILINE
    )
    if version_match is None:
        raise RuntimeError("Could not determine Docling version from pyproject.toml")

    git_result = subprocess.run(
        ["git", "rev-parse", "HEAD"],
        check=True,
        capture_output=True,
        cwd=repository_root,
        text=True,
        timeout=10,
    )
    return RuntimeInfo(
        python_version=platform.python_version(),
        platform=platform.platform(),
        docling_version=version_match.group(1),
        commit_sha=git_result.stdout.strip(),
    )


def _create_workbook(path: Path, *, merge_count: int) -> float:
    started = time.perf_counter()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "merged-ranges"

    for row in range(1, merge_count + 1):
        sheet.cell(row=row, column=1, value=f"row-{row}")
        sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=3,
        )
        sheet.cell(row=row, column=4, value=row)

    workbook.save(path)
    return time.perf_counter() - started


def _run_benchmark(
    output_dir: Path,
    *,
    merge_count: int,
    runtime_info: RuntimeInfo,
) -> BenchmarkResult:
    workbook_path = output_dir / f"merged-ranges-{merge_count}.xlsx"
    generation_seconds = _create_workbook(workbook_path, merge_count=merge_count)

    tracemalloc.start()
    started = time.perf_counter()
    converter = DocumentConverter(allowed_formats=[InputFormat.XLSX])
    document = converter.convert(workbook_path).document
    conversion_seconds = time.perf_counter() - started
    _, peak_memory = tracemalloc.get_traced_memory()
    tracemalloc.stop()

    if len(document.tables) != 1:
        raise RuntimeError(
            f"expected one table for {merge_count} merges, got {len(document.tables)}"
        )
    table = document.tables[0]
    if table.data.num_rows != merge_count or table.data.num_cols != 4:
        raise RuntimeError(
            f"unexpected table dimensions: {table.data.num_rows}x{table.data.num_cols}"
        )

    return BenchmarkResult(
        python_version=runtime_info.python_version,
        platform=runtime_info.platform,
        docling_version=runtime_info.docling_version,
        commit_sha=runtime_info.commit_sha,
        merge_count=merge_count,
        non_empty_cells=merge_count * 2,
        rectangle_area=merge_count * 4,
        generation_seconds=round(generation_seconds, 6),
        conversion_seconds=round(conversion_seconds, 6),
        peak_python_memory_mb=round(peak_memory / (1024 * 1024), 3),
        table_count=len(document.tables),
    )


def main() -> None:
    args = parse_args()
    runtime_info = _runtime_info()

    with TemporaryDirectory(prefix="docling-xlsx-merge-benchmark-") as temp_dir:
        results = [
            _run_benchmark(
                Path(temp_dir),
                merge_count=merge_count,
                runtime_info=runtime_info,
            )
            for merge_count in tqdm(
                args.merge_count,
                desc="Benchmarking merged ranges",
                unit="workbook",
            )
        ]

    output = "".join(
        f"{json.dumps(asdict(result), sort_keys=True)}\n" for result in results
    )
    sys.stdout.write(output)
    if args.output is not None:
        args.output.write_text(output, encoding="utf-8")


if __name__ == "__main__":
    main()
